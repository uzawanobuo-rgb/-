"""Core logic for syncing OTM campaign data into the CP 'おためし' sheet.

Shared by the CLI (sync_cp_otm.py) and the web tool (webtool/app.py).
"""
import csv
import re

import openpyxl
import xlrd

PCT_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*%")

# Decorative leading markers seen in CP plan names (◇◆■□●☆ etc.). OTM sometimes
# omits these, so we fall back to matching with the marker stripped from both sides.
LEADING_SYMBOLS = "◇◆■□●☆"

# プラン・キャンペーン一覧 CSV (http://magi2.atinn.jp/SimpleOutputCsv/plan) marks
# trial-campaign plan rows with these prefixes in the プラン名 column.
CAMPAIGN2_PREFIX = "【おためし入居キャンペーン②】"
CAMPAIGN1_PREFIX = "【おためし入居キャンペーン】"

SHEET_NAME_TRIAL = "おためし"

COL_D = 4    # プラン名
COL_R = 18   # おためし(現行プラン価格)
COL_Z = 26   # 賃料
COL_AF = 32  # 割引率(ルームクリーニング代設定)
COL_AG = 33  # 入居可能開始日
COL_AH = 34  # 入居可能終了日
COL_AK = 37  # 更新対象


def strip_leading_symbol(name):
    if name and name[0] in LEADING_SYMBOLS:
        return name[1:]
    return name


def to_rent_formula(raw, row):
    """'60%OFF' -> '=ROUNDUP(O{row}*(1-60%),-1)': campaign price = monthly rate(O)
    minus the OFF%, rounded up to the nearest 10 (ones digit 1-9 bumps the tens digit)."""
    if not isinstance(raw, str):
        return raw
    m = PCT_RE.match(raw)
    return f"=ROUNDUP(O{row}*(1-{m.group(1)}%),-1)" if m else raw


def to_rate_number(raw):
    """'50%OFF' -> 0.5 (real number, matches AF's existing 0% cell format)."""
    if not isinstance(raw, str):
        return raw
    m = PCT_RE.match(raw)
    return float(m.group(1)) / 100 if m else raw


def _to_date(serial, datemode):
    return xlrd.xldate.xldate_as_datetime(serial, datemode).date()


def parse_plan_campaign_csv(path):
    """Parse the 'プラン・キャンペーン一覧' CSV export.

    Returns the set of property names (raw + symbol-stripped) that have BOTH
    a '【おためし入居キャンペーン】' and a '【おためし入居キャンペーン②】' plan
    row in the プラン名 column -- i.e. it's ambiguous which campaign a request
    for that property refers to.
    """
    names1, names2 = set(), set()
    with open(path, encoding="cp932", newline="", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            plan_name = (row.get("プラン名") or "").strip()
            if plan_name.startswith(CAMPAIGN2_PREFIX):
                names2.add(plan_name[len(CAMPAIGN2_PREFIX):])
            elif plan_name.startswith(CAMPAIGN1_PREFIX):
                names1.add(plan_name[len(CAMPAIGN1_PREFIX):])

    conflict = names1 & names2
    return conflict | {strip_leading_symbol(n) for n in conflict}


def _has_campaign2_conflict(name, conflict_names):
    return name in conflict_names or strip_leading_symbol(name) in conflict_names


def parse_otm(otm_path):
    """Parse the latest (leftmost) sheet of the OTM workbook.

    Returns (sheet_name, shinki_dict, henkou_dict, skipped)
      shinki_dict: {物件名: (start_date, end_date, rent_rate_raw, clean_rate_raw)}
      henkou_dict: {物件名: (start_date, end_date)}
    """
    otm_wb = xlrd.open_workbook(otm_path)
    sheet_name = otm_wb.sheet_names()[0]
    sh = otm_wb.sheet_by_name(sheet_name)

    shinki = {}
    shinki_skipped = []
    for r in range(4, sh.nrows):
        name = sh.cell_value(r, 12)  # M列 物件名
        if not isinstance(name, str) or not name.strip():
            continue
        start = sh.cell_value(r, 13)  # N列 期間開始
        end = sh.cell_value(r, 15)    # P列 期間終了
        if not (isinstance(start, float) and isinstance(end, float)):
            shinki_skipped.append(name)
            continue
        rent_rate = sh.cell_value(r, 21)   # V列 賃料：割引率
        clean_rate = sh.cell_value(r, 22)  # W列 RC：割引率
        shinki[name] = (
            _to_date(start, otm_wb.datemode),
            _to_date(end, otm_wb.datemode),
            rent_rate,
            clean_rate,
        )

    henkou = {}
    henkou_skipped = []
    for r in range(4, sh.nrows):
        name = sh.cell_value(r, 0)  # A列 物件名
        if not isinstance(name, str) or not name.strip():
            continue
        start = sh.cell_value(r, 1)  # B列 期間開始
        end = sh.cell_value(r, 3)    # D列 期間終了
        if not (isinstance(start, float) and isinstance(end, float)):
            henkou_skipped.append(name)
            continue
        henkou[name] = (_to_date(start, otm_wb.datemode), _to_date(end, otm_wb.datemode))

    return sheet_name, shinki, henkou, {"新規依頼": shinki_skipped, "期間変更": henkou_skipped}


def _build_indexes(ws):
    exact_index = {}
    stripped_index = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=COL_D).value
        if not name:
            continue
        name = str(name)
        exact_index.setdefault(name, []).append(row)
        stripped_index.setdefault(strip_leading_symbol(name), []).append(row)
    return exact_index, stripped_index


def _resolve_row(otm_name, exact_index, stripped_index, label, ambiguous_log):
    rows = exact_index.get(otm_name)
    if rows:
        return rows[0], "exact"
    rows = stripped_index.get(strip_leading_symbol(otm_name))
    if rows and len(rows) == 1:
        return rows[0], "symbol-stripped"
    if rows and len(rows) > 1:
        ambiguous_log.append({"type": label, "name": otm_name, "candidate_rows": rows})
    return None, None


def sync(otm_path, cp_path, plan_csv_path):
    """Run the sync and return (openpyxl.Workbook, report_dict).

    otm_path / cp_path / plan_csv_path may be filesystem paths or open binary
    file objects (plan_csv_path must be a filesystem path; it's opened with a
    fixed encoding).
    """
    otm_sheet_name, shinki, henkou, skipped = parse_otm(otm_path)
    campaign2_conflicts = parse_plan_campaign_csv(plan_csv_path)

    cp_wb = openpyxl.load_workbook(cp_path, keep_vba=True, data_only=False)
    ws = cp_wb[SHEET_NAME_TRIAL]

    exact_index, stripped_index = _build_indexes(ws)

    shinki_matched, shinki_unmatched = [], []
    henkou_matched, henkou_unmatched = [], []
    ambiguous = []
    excluded_campaign2 = []

    for otm_name, (start, end) in henkou.items():
        row, how = _resolve_row(otm_name, exact_index, stripped_index, "②期間変更", ambiguous)
        if row is None:
            henkou_unmatched.append(otm_name)
            continue
        cp_name = str(ws.cell(row=row, column=COL_D).value)
        if _has_campaign2_conflict(cp_name, campaign2_conflicts):
            excluded_campaign2.append({"type": "②期間変更", "name": otm_name, "row": row, "cp_name": cp_name})
            continue
        ws.cell(row=row, column=COL_AG, value=start)
        ws.cell(row=row, column=COL_AH, value=end)
        current_price = ws.cell(row=row, column=COL_R).value
        ws.cell(row=row, column=COL_Z, value=current_price)
        ws.cell(row=row, column=COL_AK, value="○")
        henkou_matched.append({"name": otm_name, "row": row, "how": how})

    for otm_name, (start, end, rent_rate, clean_rate) in shinki.items():
        row, how = _resolve_row(otm_name, exact_index, stripped_index, "①新規依頼", ambiguous)
        if row is None:
            shinki_unmatched.append(otm_name)
            continue
        cp_name = str(ws.cell(row=row, column=COL_D).value)
        if _has_campaign2_conflict(cp_name, campaign2_conflicts):
            excluded_campaign2.append({"type": "①新規依頼", "name": otm_name, "row": row, "cp_name": cp_name})
            continue
        ws.cell(row=row, column=COL_AG, value=start)
        ws.cell(row=row, column=COL_AH, value=end)
        ws.cell(row=row, column=COL_Z, value=to_rent_formula(rent_rate, row))
        ws.cell(row=row, column=COL_AF, value=to_rate_number(clean_rate))
        ws.cell(row=row, column=COL_AK, value="○")
        shinki_matched.append({"name": otm_name, "row": row, "how": how})

    report = {
        "otm_sheet_name": otm_sheet_name,
        "skipped_rows": skipped,
        "shinki": {
            "total": len(shinki),
            "matched": shinki_matched,
            "unmatched": shinki_unmatched,
        },
        "henkou": {
            "total": len(henkou),
            "matched": henkou_matched,
            "unmatched": henkou_unmatched,
        },
        "ambiguous": ambiguous,
        "excluded_campaign2": excluded_campaign2,
    }
    return cp_wb, report

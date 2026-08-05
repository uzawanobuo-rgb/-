#!/usr/bin/env python3
"""Sync OTM campaign data into the CP 'おためし' sheet.

Reads the latest-dated sheet in the OTM workbook and writes matched
campaign period / discount info into the CP workbook's 'おためし' sheet,
using an exact-match VLOOKUP-style join on property/plan name.
"""
import re
import sys
import datetime
import xlrd
import openpyxl

PCT_RE = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*%")

# Decorative leading markers seen in CP plan names (◇◆■□●☆ etc.). OTM sometimes
# omits these, so we fall back to matching with the marker stripped from both sides.
LEADING_SYMBOLS = "◇◆■□●☆"


def strip_leading_symbol(name):
    if name and name[0] in LEADING_SYMBOLS:
        return name[1:]
    return name


def to_rent_formula(raw, row):
    """'60%OFF' -> '=N{row}*(1-60%)': campaign price = list price(N) minus the OFF%."""
    if not isinstance(raw, str):
        return raw
    m = PCT_RE.match(raw)
    return f"=N{row}*(1-{m.group(1)}%)" if m else raw


def to_rate_number(raw):
    """'50%OFF' -> 0.5 (real number, matches AF's existing 0% cell format)."""
    if not isinstance(raw, str):
        return raw
    m = PCT_RE.match(raw)
    return float(m.group(1)) / 100 if m else raw

OTM_PATH = sys.argv[1]
CP_PATH = sys.argv[2]
OUT_PATH = sys.argv[3]

# ---- 1. Load OTM (legacy .xls) and find the latest (leftmost) sheet ----
otm_wb = xlrd.open_workbook(OTM_PATH)
otm_sheet_name = otm_wb.sheet_names()[0]
sh = otm_wb.sheet_by_name(otm_sheet_name)
print(f"OTM latest sheet: {otm_sheet_name}")


def to_date(serial):
    dt = xlrd.xldate.xldate_as_datetime(serial, otm_wb.datemode)
    return dt.date()


# ---- ①新規依頼 block: M=物件名(12), N=期間開始(13), P=期間終了(15), V=賃料割引率(21), W=RC割引率(22) ----
shinki = {}
shinki_skipped = []
for r in range(4, sh.nrows):
    name = sh.cell_value(r, 12)
    if not isinstance(name, str) or not name.strip():
        continue
    start = sh.cell_value(r, 13)
    end = sh.cell_value(r, 15)
    if not (isinstance(start, float) and isinstance(end, float)):
        shinki_skipped.append(name)
        continue
    rent_rate = sh.cell_value(r, 21)
    clean_rate = sh.cell_value(r, 22)
    shinki[name] = (to_date(start), to_date(end), rent_rate, clean_rate)

# ---- ②期間変更 block: A=物件名(0), B=期間開始(1), D=期間終了(3) ----
henkou = {}
henkou_skipped = []
for r in range(4, sh.nrows):
    name = sh.cell_value(r, 0)
    if not isinstance(name, str) or not name.strip():
        continue
    start = sh.cell_value(r, 1)
    end = sh.cell_value(r, 3)
    if not (isinstance(start, float) and isinstance(end, float)):
        henkou_skipped.append(name)
        continue
    henkou[name] = (to_date(start), to_date(end))

print(f"①新規依頼 candidates: {len(shinki)} (skipped/no-date rows: {shinki_skipped})")
print(f"②期間変更 candidates: {len(henkou)} (skipped/no-date rows: {henkou_skipped})")

# ---- 2. Load CP workbook (.xlsm, keep macros) ----
cp_wb = openpyxl.load_workbook(CP_PATH, keep_vba=True, data_only=False)
ws = cp_wb["おためし"]

COL_D = 4    # プラン名
COL_R = 18   # おためし(現行プラン価格)
COL_Z = 26   # 賃料
COL_AF = 32  # 割引率(ルームクリーニング代設定)
COL_AG = 33  # 入居可能開始日
COL_AH = 34  # 入居可能終了日
COL_AK = 37  # 更新対象

# ---- Build CP name -> row lookups (exact, and symbol-stripped fallback) ----
exact_index = {}
stripped_index = {}
for row in range(4, ws.max_row + 1):
    name = ws.cell(row=row, column=COL_D).value
    if not name:
        continue
    name = str(name)
    exact_index.setdefault(name, []).append(row)
    stripped_index.setdefault(strip_leading_symbol(name), []).append(row)


def resolve_row(otm_name, label, ambiguous_log):
    """Exact match first; fall back to symbol-stripped match if unambiguous."""
    rows = exact_index.get(otm_name)
    if rows:
        return rows[0], "exact"
    rows = stripped_index.get(strip_leading_symbol(otm_name))
    if rows and len(rows) == 1:
        return rows[0], "symbol-stripped"
    if rows and len(rows) > 1:
        ambiguous_log.append((label, otm_name, rows))
    return None, None


shinki_matched = []
shinki_unmatched = []
henkou_matched = []
henkou_unmatched = []
ambiguous = []

for otm_name, (start, end) in henkou.items():
    row, how = resolve_row(otm_name, "②期間変更", ambiguous)
    if row is None:
        henkou_unmatched.append(otm_name)
        continue
    ws.cell(row=row, column=COL_AG, value=start)
    ws.cell(row=row, column=COL_AH, value=end)
    current_price = ws.cell(row=row, column=COL_R).value
    ws.cell(row=row, column=COL_Z, value=current_price)
    ws.cell(row=row, column=COL_AK, value="○")
    henkou_matched.append((otm_name, row, how))

for otm_name, (start, end, rent_rate, clean_rate) in shinki.items():
    row, how = resolve_row(otm_name, "①新規依頼", ambiguous)
    if row is None:
        shinki_unmatched.append(otm_name)
        continue
    ws.cell(row=row, column=COL_AG, value=start)
    ws.cell(row=row, column=COL_AH, value=end)
    ws.cell(row=row, column=COL_Z, value=to_rent_formula(rent_rate, row))
    ws.cell(row=row, column=COL_AF, value=to_rate_number(clean_rate))
    ws.cell(row=row, column=COL_AK, value="○")
    shinki_matched.append((otm_name, row, how))

cp_wb.save(OUT_PATH)

print(f"\n①新規依頼 matched: {len(shinki_matched)} / {len(shinki)}")
for n, r, how in shinki_matched:
    if how != "exact":
        print(f"  [{how}] {n!r} -> row {r}")
print(f"①新規依頼 unmatched: {shinki_unmatched}")
print(f"②期間変更 matched: {len(henkou_matched)} / {len(henkou)}")
for n, r, how in henkou_matched:
    if how != "exact":
        print(f"  [{how}] {n!r} -> row {r}")
print(f"②期間変更 unmatched: {henkou_unmatched}")
if ambiguous:
    print("\nAMBIGUOUS (multiple CP rows share the symbol-stripped name; skipped, needs manual review):")
    for label, n, rows in ambiguous:
        print(f"  [{label}] {n!r} -> candidate rows {rows}")
print(f"\nSaved: {OUT_PATH}")

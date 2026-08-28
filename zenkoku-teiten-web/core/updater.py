# -*- coding: utf-8 -*-
"""
全国定点.xlsx 更新ロジック(コア部分)

update_zenkoku_teiten.py (CLI版) と app.py (Streamlit版) の両方から
このモジュールの関数を呼び出す。ロジック自体はCLI版から変更していない。
"""

import re
import io
import datetime
import openpyxl

FIXEDPOINT_SHEET = "FixedPoint"

COL = {
    "day_serial": 2,
    "rate_nat": 4, "rate_nat_delta": 5,
    "rate_tky": 6, "rate_tky_delta": 7,
    "rate_nag": 8, "rate_nag_delta": 9,
    "rate_osa": 10, "rate_osa_delta": 11,
    "sales_nat": 12,
    "sales_tky": 15,
    "sales_nag": 18,
    "sales_osa": 21,
    "sales_sht": 24,
    "rooms_nat": 26,
    "rooms_tky": 29,
    "rooms_nag": 32,
    "rooms_osa": 35,
    "rooms_sht": 38,
}

BLOCK_LABELS = ["当月", "1か月後", "2か月後", "3か月後"]

AREA_LABELS = {
    "nag": "名古屋WM",
    "tky": "東京WM",
    "osa": "大阪WM",
    "sht": "社宅・賃貸",
}


class UpdateError(Exception):
    pass


def parse_date_from_filename(filename: str) -> datetime.date:
    m = re.search(r"FixedPoint(\d{4})(\d{2})(\d{2})\d{6}", filename)
    if not m:
        raise UpdateError(f"ファイル名から日付を抽出できません: {filename}")
    y, mo, d = map(int, m.groups())
    return datetime.date(y, mo, d)


def _add_months(d: datetime.date, n: int) -> datetime.date:
    month = d.month - 1 + n
    year = d.year + month // 12
    month = month % 12 + 1
    return datetime.date(year, month, 1)


def find_area_rows(ws_raw):
    rows = {}
    max_row = min(ws_raw.max_row, 60)
    for r in range(1, max_row + 1):
        v = ws_raw.cell(row=r, column=15).value
        if isinstance(v, str):
            for key, label in AREA_LABELS.items():
                if v.strip() == label:
                    rows[key] = r
    missing = set(AREA_LABELS) - set(rows)
    if missing:
        raise UpdateError(f"rawファイル内でエリアラベルが見つかりません: {missing}")
    return rows


def find_month_base_col(ws_raw, target_date: datetime.date):
    label = f"{target_date.year}/{target_date.month:02d}"
    max_col = ws_raw.max_column
    for c in range(1, max_col + 1):
        if ws_raw.cell(row=2, column=c).value == label:
            return c
    return None


def extract_area_values(ws_raw, label_row, sales_row, month_col):
    occupied = ws_raw.cell(row=label_row, column=month_col).value or 0
    total = ws_raw.cell(row=label_row, column=month_col + 2).value or 0
    sales_gross = ws_raw.cell(row=sales_row, column=month_col).value or 0
    discount = ws_raw.cell(row=sales_row, column=month_col + 2).value or 0
    sales_net = sales_gross + discount
    return occupied, total, sales_net


def find_block_start_row(ws_target, label):
    for r in range(1, ws_target.max_row + 1):
        if ws_target.cell(row=r, column=1).value == label:
            return r
    raise UpdateError(f"全国定点シートにブロック '{label}' が見つかりません")


def process_one_file(wb_target_write, wb_target_read, raw_filename, raw_bytes, log=None):
    """
    wb_target_write: 書き込み先ワークブック (data_only=False で開いたもの)
    wb_target_read : 前日値参照用ワークブック (data_only=True で開いたもの。
                      直前の保存内容を読み直したもの)
    raw_filename   : rawファイルのファイル名(日付抽出に使用)
    raw_bytes      : rawファイルのバイト列 (bytes) または ファイルパス(str)
    log            : リストを渡すとログ行が追記される
    """
    if log is None:
        log = []

    raw_date = parse_date_from_filename(raw_filename)
    sheet_name = f"{raw_date.year}年{raw_date.month}月"

    if sheet_name not in wb_target_write.sheetnames:
        raise UpdateError(
            f"全国定点.xlsx に '{sheet_name}' シートがありません。"
            f"月が変わる場合は事前にテンプレートをコピーしてシートを用意してください。"
        )

    ws_target_w = wb_target_write[sheet_name]
    ws_target_r = wb_target_read[sheet_name]

    wb_raw = openpyxl.load_workbook(
        io.BytesIO(raw_bytes) if isinstance(raw_bytes, (bytes, bytearray)) else raw_bytes,
        data_only=True, read_only=True,
    )
    ws_raw = wb_raw[FIXEDPOINT_SHEET]

    area_rows = find_area_rows(ws_raw)
    day_of_month = raw_date.day

    for block_idx, block_label in enumerate(BLOCK_LABELS):
        target_month_date = _add_months(raw_date, block_idx)
        base_col = find_month_base_col(ws_raw, target_month_date)
        if base_col is None:
            continue  # rawデータがまだその先の月まで届いていない

        vals = {}
        for key in ("nag", "tky", "osa", "sht"):
            label_row = area_rows[key]
            sales_row = label_row + 1
            occ, tot, sales_net = extract_area_values(ws_raw, label_row, sales_row, base_col)
            vals[key] = {"occ": occ, "tot": tot, "sales": sales_net}

        rate = {}
        for key in ("nag", "tky", "osa"):
            occ, tot = vals[key]["occ"], vals[key]["tot"]
            rate[key] = (occ / tot * 100) if tot else 0

        occ_sum = sum(vals[k]["occ"] for k in ("nag", "tky", "osa", "sht"))
        tot_sum = sum(vals[k]["tot"] for k in ("nag", "tky", "osa", "sht"))
        rate["nat"] = (occ_sum / tot_sum * 100) if tot_sum else 0

        block_start_row = find_block_start_row(ws_target_w, block_label)
        target_row = block_start_row + (day_of_month - 1)

        rate_col_map = {
            "nat": (COL["rate_nat"], COL["rate_nat_delta"]),
            "tky": (COL["rate_tky"], COL["rate_tky_delta"]),
            "nag": (COL["rate_nag"], COL["rate_nag_delta"]),
            "osa": (COL["rate_osa"], COL["rate_osa_delta"]),
        }
        for key, (abs_col, delta_col) in rate_col_map.items():
            new_val = round(rate[key], 2)
            prev_val = ws_target_r.cell(row=target_row - 1, column=abs_col).value
            ws_target_w.cell(row=target_row, column=abs_col, value=new_val)
            if isinstance(prev_val, (int, float)):
                ws_target_w.cell(row=target_row, column=delta_col, value=round(new_val - prev_val, 2))

        sales_col_map = {
            "tky": COL["sales_tky"], "nag": COL["sales_nag"],
            "osa": COL["sales_osa"], "sht": COL["sales_sht"],
        }
        for key, col in sales_col_map.items():
            ws_target_w.cell(row=target_row, column=col, value=vals[key]["sales"])

        rooms_col_map = {
            "tky": COL["rooms_tky"], "nag": COL["rooms_nag"],
            "osa": COL["rooms_osa"], "sht": COL["rooms_sht"],
        }
        for key, col in rooms_col_map.items():
            ws_target_w.cell(row=target_row, column=col, value=vals[key]["occ"])

        log.append(
            f"[{raw_filename}] {sheet_name} / {block_label} row={target_row} / "
            f"稼働率(全国 {rate['nat']:.2f}% 東京 {rate['tky']:.2f}% "
            f"名古屋 {rate['nag']:.2f}% 大阪 {rate['osa']:.2f}%)"
        )

    return log


def update_master(master_bytes, raw_files):
    """
    master_bytes: 全国定点.xlsx のバイト列
    raw_files   : [(filename, bytes), ...] のリスト(日付順でなくてもよい)
    戻り値      : (更新済みワークブックのバイト列, ログ行のリスト)
    """
    raw_files_sorted = sorted(raw_files, key=lambda x: parse_date_from_filename(x[0]))

    wb_write = openpyxl.load_workbook(io.BytesIO(master_bytes), data_only=False)
    full_log = []

    for filename, raw_bytes in raw_files_sorted:
        # 直前までの更新結果を読み直して「前日値」参照に使う
        buf = io.BytesIO()
        wb_write.save(buf)
        buf.seek(0)
        wb_read = openpyxl.load_workbook(buf, data_only=True)

        process_one_file(wb_write, wb_read, filename, raw_bytes, log=full_log)

    out_buf = io.BytesIO()
    wb_write.save(out_buf)
    return out_buf.getvalue(), full_log
